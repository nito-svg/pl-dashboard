import streamlit as st
import pdfplumber
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import re
from io import BytesIO
from datetime import datetime

st.set_page_config(
    page_title="月次試算表ダッシュボード",
    page_icon="◐",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== カスタムCSS（ミニマル&クリーン）=====
st.markdown("""
<style>
    /* メインタイトルとヘッダー */
    h1 {
        font-weight: 300 !important;
        letter-spacing: -0.02em;
        color: #1a1a1a;
    }
    h2 {
        font-weight: 400 !important;
        color: #1a1a1a;
        margin-top: 2rem !important;
        letter-spacing: -0.01em;
    }
    h3 {
        font-weight: 400 !important;
        color: #4a4a4a;
        font-size: 1.1rem !important;
    }
    
    /* メトリックカード */
    [data-testid="stMetricValue"] {
        font-weight: 400 !important;
        font-size: 1.5rem !important;
        color: #1a1a1a;
    }
    [data-testid="stMetricLabel"] {
        font-size: 0.75rem !important;
        color: #888 !important;
        letter-spacing: 0.05em;
        text-transform: uppercase;
    }
    [data-testid="stMetricDelta"] {
        font-size: 0.8rem !important;
    }
    
    /* セクション間の余白 */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        max-width: 1400px;
    }
    
    /* ボタン */
    .stDownloadButton > button {
        background: #1a1a1a;
        color: white;
        border: none;
        border-radius: 4px;
        padding: 0.6rem 1.5rem;
        font-weight: 400;
        letter-spacing: 0.02em;
    }
    .stDownloadButton > button:hover {
        background: #333;
    }
    
    /* 区切り線 */
    hr {
        margin: 2rem 0 !important;
        border-color: #eaeaea !important;
    }
    
    /* サイドバー */
    [data-testid="stSidebar"] {
        background: #fafafa;
    }
    
    /* タブ */
    [data-baseweb="tab-list"] {
        gap: 2rem;
    }
    
    /* キャプション */
    .small-caption {
        color: #888;
        font-size: 0.75rem;
        letter-spacing: 0.05em;
        text-transform: uppercase;
        margin-bottom: 0.5rem;
    }
    
    /* 数字大きい表示 */
    .big-number {
        font-size: 2rem;
        font-weight: 300;
        color: #1a1a1a;
        letter-spacing: -0.02em;
        margin: 0;
    }
</style>
""", unsafe_allow_html=True)

# ===== 重要項目の定義 =====
MAIN_KEY_ITEMS = [
    "売上高合計", "売上原価", "売上総利益",
    "販売費及び一般管理費合計", "営業利益",
    "営業外収益合計", "営業外費用合計", "経常利益",
    "特別利益合計", "特別損失合計",
    "税引前当期純利益", "当期純利益"
]

SGA_KEY_ITEMS = [
    "役員報酬", "給料手当", "賞与", "法定福利費", "福利厚生費",
    "地代家賃", "減価償却費",
    "広告宣伝費", "販売促進費", "交際費", "旅費交通費",
    "通信費", "水道光熱費", "消耗品費",
    "支払手数料", "業務委託費", "支払報酬料"
]


def clean_name(item):
    return item.replace("【", "").replace("】", "").strip()


def extract_all_data_from_pdf(pdf_file):
    months = []
    all_items_data = {}
    
    with pdfplumber.open(pdf_file) as pdf:
        all_text = ""
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                all_text += page_text + "\n"
    
    lines = all_text.split("\n")
    
    pl_start = -1
    for i, line in enumerate(lines):
        if "損益計算書" in line:
            pl_start = i
            break
    
    if pl_start == -1:
        return [], {}, False
    
    pl_lines = lines[pl_start:]
    
    month_pattern = re.compile(r'(\d+月)')
    has_total = False
    for line in pl_lines:
        matches = month_pattern.findall(line)
        if len(matches) >= 3:
            seen = []
            for m in matches:
                if m not in seen:
                    seen.append(m)
            if len(seen) >= 3:
                months = seen
                if "合計" in line:
                    has_total = True
                break
    
    if not months:
        return [], {}, False
    
    expected_cols = len(months) + (1 if has_total else 0)
    
    for i, line in enumerate(pl_lines):
        stripped = line.strip()
        if not stripped:
            continue
        
        numbers = re.findall(r'-?[\d,]+', stripped)
        numbers = [n for n in numbers if re.search(r'\d', n)]
        
        if len(numbers) >= expected_cols:
            first_num_pos = -1
            for num in numbers:
                pos = stripped.find(num)
                if pos > 0:
                    first_num_pos = pos
                    break
            
            if first_num_pos <= 0:
                continue
            
            item_name = stripped[:first_num_pos].strip()
            clean_item = clean_name(item_name)
            
            if not clean_item or len(clean_item) > 30:
                continue
            
            values = []
            try:
                for j in range(len(months)):
                    val = int(numbers[j].replace(",", ""))
                    values.append(val)
                all_items_data[clean_item] = values
            except (ValueError, IndexError):
                continue
    
    return months, all_items_data, has_total


def categorize_items(all_items):
    main_items = []
    sga_items = []
    other_items = []
    
    for item in all_items.keys():
        if item in MAIN_KEY_ITEMS:
            main_items.append(item)
        elif item in SGA_KEY_ITEMS:
            sga_items.append(item)
        else:
            other_items.append(item)
    
    return main_items, sga_items, other_items


def create_excel_file(months, all_data, main_items, sga_items, other_items):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="EAEAEA"),
        right=Side(style="thin", color="EAEAEA"),
        top=Side(style="thin", color="EAEAEA"),
        bottom=Side(style="thin", color="EAEAEA"),
    )
    
    def write_sheet(sheet_name, items):
        if not items:
            return
        ws = wb.create_sheet(sheet_name)
        headers = ["項目"] + months + ["合計"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item).border = thin_border
            values = all_data.get(item, [0] * len(months))
            for col_idx, val in enumerate(values, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.number_format = '¥#,##0;[Red]¥-#,##0'
            total_cell = ws.cell(row=row_idx, column=len(months) + 2, value=sum(values))
            total_cell.border = thin_border
            total_cell.number_format = '¥#,##0;[Red]¥-#,##0'
            total_cell.font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(months) + 3):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16
    
    write_sheet("主要指標", main_items)
    write_sheet("販管費", sga_items)
    write_sheet("その他", other_items)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def format_amount(amount):
    """金額を読みやすくフォーマット"""
    if abs(amount) >= 100000000:  # 1億以上
        return f"¥{amount/100000000:.2f}億"
    elif abs(amount) >= 10000:  # 1万以上
        return f"¥{amount/10000:,.0f}万"
    else:
        return f"¥{amount:,}"


# プロットリーのテーマ設定
PLOTLY_THEME = {
    "layout": {
        "font": {"family": "system-ui, -apple-system, sans-serif", "size": 12, "color": "#1a1a1a"},
        "plot_bgcolor": "white",
        "paper_bgcolor": "white",
        "colorway": ["#1a1a1a", "#4a90e2", "#7b8ba0", "#c9a96e", "#82a47f", "#b86b6b", "#5d6d7e", "#a08fb3"],
        "xaxis": {"gridcolor": "#f0f0f0", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}},
        "yaxis": {"gridcolor": "#f0f0f0", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}},
    }
}


# ===== UI =====
st.title("月次試算表ダッシュボード")
st.markdown(
    "<p style='color: #888; font-size: 0.9rem; margin-top: -0.5rem; margin-bottom: 2rem;'>"
    "Monthly Trial Balance Dashboard"
    "</p>",
    unsafe_allow_html=True
)

uploaded_file = st.file_uploader("PDFファイルをアップロード", type="pdf", label_visibility="visible")

if uploaded_file is not None:
    months, all_data, has_total = extract_all_data_from_pdf(uploaded_file)
    
    if not months or not all_data:
        st.error("データを抽出できませんでした。PDFのフォーマットをご確認ください。")
    else:
        main_items, sga_items, other_items = categorize_items(all_data)
        
        # ファイル情報を控えめに
        st.markdown(
            f"<div style='background: #fafafa; padding: 0.75rem 1rem; border-radius: 4px; "
            f"border: 1px solid #eaeaea; margin-bottom: 2rem; font-size: 0.85rem; color: #666;'>"
            f"<strong style='color: #1a1a1a;'>{uploaded_file.name}</strong>　"
            f"／　期間：{months[0]} 〜 {months[-1]}　"
            f"／　検出項目：主要 {len(main_items)} ・ 販管費 {len(sga_items)} ・ その他 {len(other_items)}"
            f"</div>",
            unsafe_allow_html=True
        )
        
        # ===== サイドバー =====
        st.sidebar.markdown("### 表示設定")
        
        display_mode = st.sidebar.radio(
            "表示モード",
            ["重要項目のみ", "全項目", "カスタム選択"],
            label_visibility="collapsed"
        )
        
        if display_mode == "重要項目のみ":
            displayed_main = main_items
            displayed_sga = sga_items
        elif display_mode == "全項目":
            displayed_main = main_items
            displayed_sga = sga_items + other_items
        else:
            st.sidebar.markdown("**主要指標**")
            displayed_main = st.sidebar.multiselect(
                "選択", main_items, default=main_items,
                key="main_select", label_visibility="collapsed"
            )
            st.sidebar.markdown("**販管費・その他**")
            all_other = sga_items + other_items
            displayed_sga = st.sidebar.multiselect(
                "選択", all_other, default=sga_items,
                key="sga_select", label_visibility="collapsed"
            )
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("### 警告閾値")
        warning_threshold = st.sidebar.slider(
            "変動率（%）", 5, 100, 20, 5, label_visibility="collapsed"
        )
        st.sidebar.caption(f"前月比 ±{warning_threshold}% を超えた項目を強調")
        
        # ===== 主要指標サマリー =====
        if displayed_main:
            st.markdown("## 主要指標")
            st.markdown(
                f"<p style='color: #888; font-size: 0.8rem; margin-top: -0.5rem; "
                f"letter-spacing: 0.05em;'>{months[-1]}（前月比）</p>",
                unsafe_allow_html=True
            )
            
            latest_idx = len(months) - 1
            
            for row_start in range(0, len(displayed_main), 4):
                row_items = displayed_main[row_start:row_start + 4]
                cols = st.columns(4)
                for i, item in enumerate(row_items):
                    with cols[i]:
                        values = all_data[item]
                        latest_val = values[latest_idx]
                        if latest_idx > 0 and values[latest_idx - 1] != 0:
                            prev_val = values[latest_idx - 1]
                            rate = (latest_val - prev_val) / abs(prev_val) * 100
                            st.metric(
                                label=item,
                                value=format_amount(latest_val),
                                delta=f"{rate:+.1f}%"
                            )
                        else:
                            st.metric(
                                label=item,
                                value=format_amount(latest_val)
                            )
            
            # 推移グラフ
            st.markdown("## 推移")
            df_main = pd.DataFrame({item: all_data[item] for item in displayed_main}, index=months)
            df_main.index.name = "月"
            df_main_long = df_main.reset_index().melt(
                id_vars="月", var_name="項目", value_name="金額"
            )
            fig_main = px.line(
                df_main_long, x="月", y="金額", color="項目",
                markers=True,
                color_discrete_sequence=["#1a1a1a", "#4a90e2", "#7b8ba0", "#c9a96e", "#82a47f", "#b86b6b", "#5d6d7e", "#a08fb3", "#6b8ca8", "#a5907a", "#7a9181", "#9c7474"]
            )
            fig_main.update_layout(
                height=450,
                hovermode="x unified",
                font={"family": "system-ui, -apple-system, sans-serif", "size": 12, "color": "#1a1a1a"},
                plot_bgcolor="white",
                paper_bgcolor="white",
                xaxis={"gridcolor": "#f5f5f5", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}, "title": ""},
                yaxis={"gridcolor": "#f5f5f5", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}, "title": ""},
                legend={"orientation": "v", "yanchor": "top", "y": 1, "xanchor": "left", "x": 1.02},
                margin={"l": 60, "r": 20, "t": 20, "b": 40}
            )
            fig_main.update_traces(line={"width": 2})
            st.plotly_chart(fig_main, use_container_width=True)
            
            # 期間累計
            st.markdown("## 期間累計")
            st.markdown(
                f"<p style='color: #888; font-size: 0.8rem; margin-top: -0.5rem; "
                f"letter-spacing: 0.05em;'>{months[0]} 〜 {months[-1]}</p>",
                unsafe_allow_html=True
            )
            for row_start in range(0, len(displayed_main), 4):
                row_items = displayed_main[row_start:row_start + 4]
                cols = st.columns(4)
                for i, item in enumerate(row_items):
                    with cols[i]:
                        total = sum(all_data[item])
                        st.metric(label=item, value=f"¥{total:,}")
        
        # ===== 販管費 =====
        if displayed_sga:
            st.markdown("---")
            st.markdown("## 販管費 分析")
            
            if len(months) >= 2:
                st.markdown(
                    f"<p style='color: #888; font-size: 0.8rem; margin-top: -0.5rem; "
                    f"letter-spacing: 0.05em;'>前月比　{months[-2]} → {months[-1]}</p>",
                    unsafe_allow_html=True
                )
                
                for row_start in range(0, len(displayed_sga), 4):
                    row_items = displayed_sga[row_start:row_start + 4]
                    cols = st.columns(4)
                    for i, item in enumerate(row_items):
                        with cols[i]:
                            values = all_data[item]
                            latest_val = values[-1]
                            prev_val = values[-2]
                            if prev_val != 0:
                                rate = (latest_val - prev_val) / abs(prev_val) * 100
                                st.metric(
                                    label=item,
                                    value=format_amount(latest_val),
                                    delta=f"{rate:+.1f}%",
                                    delta_color="inverse"
                                )
                            else:
                                st.metric(label=item, value=format_amount(latest_val))
            
            # 推移グラフ
            st.markdown("### 推移")
            df_sga = pd.DataFrame({item: all_data[item] for item in displayed_sga}, index=months)
            df_sga.index.name = "月"
            df_sga_long = df_sga.reset_index().melt(
                id_vars="月", var_name="項目", value_name="金額"
            )
            fig_sga = px.line(
                df_sga_long, x="月", y="金額", color="項目", markers=True
            )
            fig_sga.update_layout(
                height=500,
                hovermode="x unified",
                font={"family": "system-ui, -apple-system, sans-serif", "size": 12, "color": "#1a1a1a"},
                plot_bgcolor="white",
                paper_bgcolor="white",
                xaxis={"gridcolor": "#f5f5f5", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}, "title": ""},
                yaxis={"gridcolor": "#f5f5f5", "linecolor": "#e0e0e0", "tickfont": {"color": "#666"}, "title": ""},
                legend={"orientation": "v", "yanchor": "top", "y": 1, "xanchor": "left", "x": 1.02},
                margin={"l": 60, "r": 20, "t": 20, "b": 40}
            )
            fig_sga.update_traces(line={"width": 1.5})
            st.plotly_chart(fig_sga, use_container_width=True)
            
            # 大きく変動した項目
            if len(months) >= 2:
                st.markdown("### 変動分析")
                changes = []
                for item in displayed_sga:
                    values = all_data[item]
                    latest_val = values[-1]
                    prev_val = values[-2]
                    if prev_val != 0:
                        diff = latest_val - prev_val
                        rate = round(diff / abs(prev_val) * 100, 1)
                        changes.append({
                            "項目": item,
                            "前月": prev_val,
                            "当月": latest_val,
                            "差額": diff,
                            "変動率(%)": rate
                        })
                
                if changes:
                    df_changes = pd.DataFrame(changes)
                    df_changes["abs_rate"] = df_changes["変動率(%)"].abs()
                    df_changes = df_changes.sort_values("abs_rate", ascending=False).drop(columns=["abs_rate"])
                    
                    df_changes_display = df_changes.copy()
                    df_changes_display["前月"] = df_changes_display["前月"].apply(lambda x: f"¥{x:,}")
                    df_changes_display["当月"] = df_changes_display["当月"].apply(lambda x: f"¥{x:,}")
                    df_changes_display["差額"] = df_changes_display["差額"].apply(
                        lambda x: f"+¥{x:,}" if x > 0 else f"-¥{abs(x):,}" if x < 0 else "¥0"
                    )
                    df_changes_display["変動率(%)"] = df_changes_display["変動率(%)"].apply(
                        lambda x: f"+{x}%" if x > 0 else f"{x}%"
                    )
                    
                    st.dataframe(df_changes_display, hide_index=True, use_container_width=True)
                    
                    big_changes = df_changes[df_changes["変動率(%)"].abs() >= warning_threshold]
                    if not big_changes.empty:
                        warning_html = f"""
                        <div style='background: #fff8e7; border-left: 3px solid #d4a843; 
                                    padding: 1rem 1.25rem; border-radius: 4px; margin: 1rem 0;'>
                            <p style='color: #1a1a1a; font-size: 0.85rem; font-weight: 500; 
                                      margin: 0 0 0.5rem; letter-spacing: 0.02em;'>
                                ±{warning_threshold}% 以上の変動 — {len(big_changes)}件
                            </p>
                            <ul style='margin: 0; padding-left: 1.25rem; color: #4a4a4a; 
                                       font-size: 0.85rem; line-height: 1.8;'>
                        """
                        for _, row in big_changes.iterrows():
                            warning_html += (
                                f"<li>{row['項目']}　{row['変動率(%)']:+.1f}%　"
                                f"<span style='color: #888;'>"
                                f"¥{row['前月']:,} → ¥{row['当月']:,}</span></li>"
                            )
                        warning_html += "</ul></div>"
                        st.markdown(warning_html, unsafe_allow_html=True)
        
        # ===== Excel出力 =====
        st.markdown("---")
        st.markdown("## ダウンロード")
        col1, col2 = st.columns([1, 3])
        with col1:
            excel_data = create_excel_file(months, all_data, main_items, sga_items, other_items)
            filename = f"月次試算表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="Excelファイルでダウンロード",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col2:
            st.markdown(
                f"<p style='color: #888; font-size: 0.8rem; padding-top: 0.6rem;'>"
                f"全 {len(main_items) + len(sga_items) + len(other_items)} 項目を3シート構成で出力します"
                f"</p>",
                unsafe_allow_html=True
            )
        
        # ===== 全データ一覧 =====
        st.markdown("---")
        with st.expander("全データを表示"):
            tab1, tab2, tab3 = st.tabs([f"主要指標 ({len(main_items)})", f"販管費 ({len(sga_items)})", f"その他 ({len(other_items)})"])
            
            with tab1:
                if main_items:
                    df_show = pd.DataFrame({item: all_data[item] for item in main_items}, index=months).T
                    df_show.columns = months
                    df_show_display = df_show.copy()
                    for col in df_show_display.columns:
                        df_show_display[col] = df_show_display[col].apply(lambda x: f"¥{x:,}")
                    st.dataframe(df_show_display, use_container_width=True)
            
            with tab2:
                if sga_items:
                    df_show = pd.DataFrame({item: all_data[item] for item in sga_items}, index=months).T
                    df_show.columns = months
                    df_show_display = df_show.copy()
                    for col in df_show_display.columns:
                        df_show_display[col] = df_show_display[col].apply(lambda x: f"¥{x:,}")
                    st.dataframe(df_show_display, use_container_width=True)
            
            with tab3:
                if other_items:
                    df_show = pd.DataFrame({item: all_data[item] for item in other_items}, index=months).T
                    df_show.columns = months
                    df_show_display = df_show.copy()
                    for col in df_show_display.columns:
                        df_show_display[col] = df_show_display[col].apply(lambda x: f"¥{x:,}")
                    st.dataframe(df_show_display, use_container_width=True)

else:
    # 初期画面（ファイル未アップロード時）
    st.markdown(
        "<div style='text-align: center; padding: 4rem 2rem; color: #888;'>"
        "<p style='font-size: 0.9rem; letter-spacing: 0.05em;'>"
        "PDFファイルをアップロードして開始してください"
        "</p></div>",
        unsafe_allow_html=True
    )